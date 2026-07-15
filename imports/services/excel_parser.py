"""Safe spreadsheet parsing.

openpyxl runs with read_only=True and data_only=True: formulas are replaced by
their last cached values and macros are never executed. Rep strings like
"8–12", "10/leg", "AMRAP", or "30 seconds" are parsed defensively — anything
unrecognized is kept as text rather than rejected.
"""
import csv
import io
import re

from django.conf import settings

MAPPABLE_FIELDS = [
    ("week", "Week"),
    ("day", "Day"),
    ("weekday", "Weekday"),
    ("workout_name", "Workout name"),
    ("exercise", "Exercise"),
    ("sets", "Sets"),
    ("reps", "Reps"),
    ("weight", "Weight"),
    ("percentage", "Percentage"),
    ("rir", "RIR"),
    ("rpe", "RPE"),
    ("rest", "Rest"),
    ("tempo", "Tempo"),
    ("notes", "Notes"),
    ("superset", "Superset"),
]

MAX_ROWS = 2000
MAX_COLS = 40
PREVIEW_ROWS = 8


class WorkbookLimitError(ValueError):
    """Raised when a workbook is valid but too large to interpret safely."""


# Common labels coaches use in Excel/CSV training templates. Suggestions stay
# fully editable in the mapper, so a guess never becomes a hidden rule.
COLUMN_ALIASES = {
    "week": {"week", "week number", "wk"},
    "day": {"day", "day number", "training day", "workout day"},
    "weekday": {"weekday", "day of week", "training weekday", "schedule day"},
    "workout_name": {"workout", "workout name", "session", "session name", "split"},
    "exercise": {"exercise", "exercise name", "movement", "lift"},
    "sets": {"sets", "set", "number of sets", "set count"},
    "reps": {
        "reps", "rep", "target reps", "repetitions", "rep range",
        "sets x reps", "set rep scheme", "scheme", "prescription",
    },
    "weight": {
        "weight", "load", "weight lb", "load lb", "working weight",
        "actual weight", "target load",
    },
    "percentage": {"percentage", "percent", "%", "% of 1rm", "percent of 1rm"},
    "rir": {"rir", "reps in reserve"},
    "rpe": {"rpe", "rate of perceived exertion"},
    "rest": {"rest", "rest seconds", "rest time", "rest period"},
    "tempo": {"tempo", "cadence"},
    "notes": {"notes", "note", "instructions", "intensity notes", "coach notes"},
    "superset": {"superset", "superset group", "group"},
}


def suggest_column_mapping(header):
    """Return a best-effort map of import fields to spreadsheet columns."""
    suggestions = {}
    for index, value in enumerate(header):
        label = re.sub(r"[^a-z0-9%]+", " ", str(value or "").lower()).strip()
        if not label:
            continue
        # "Set 1", "Set 2", etc. are per-set load columns, not a count of
        # how many sets to perform. The smart parser handles those together.
        if re.fullmatch(r"set\s*\d+", label):
            continue
        if (
            ("set" in label and "rep" in label)
            or label in {"scheme", "prescription"}
        ):
            suggestions.setdefault("reps", index)
            continue
        if label in COLUMN_ALIASES["weekday"]:
            suggestions.setdefault("weekday", index)
            continue
        for field, aliases in COLUMN_ALIASES.items():
            if field in suggestions:
                continue
            if label in aliases or any(alias in label for alias in aliases if len(alias) > 2):
                suggestions[field] = index
                break
    return suggestions


def detect_header_row(rows, search_limit=50):
    """Find the most likely column-header row in a title-heavy worksheet."""
    best_index = 0
    best_mapping = suggest_column_mapping(rows[0]) if rows else {}
    best_score = -1
    for index, row in enumerate(rows[:search_limit]):
        mapping = suggest_column_mapping(row)
        if "exercise" not in mapping:
            continue
        score = len(mapping) + 3
        if score > best_score:
            best_index = index
            best_mapping = mapping
            best_score = score
    return best_index, best_mapping


def _normalized_label(value):
    return re.sub(r"[^a-z0-9%]+", " ", str(value or "").lower()).strip()


def _week_number(sheet_name, rows):
    """Find a week number from a worksheet name or its opening title rows."""
    match = re.match(r"^\s*week\s*(\d{1,2})\b", str(sheet_name), re.IGNORECASE)
    if match and 1 <= int(match.group(1)) <= 52:
        return int(match.group(1))
    for candidate in (cell for row in rows[:3] for cell in row if cell):
        match = re.match(r"^\s*week\s*(\d{1,2})\b", str(candidate), re.IGNORECASE)
        if match and 1 <= int(match.group(1)) <= 52:
            return int(match.group(1))
    return None


def _weekday_number(value):
    text = str(value or "").strip().lower()
    names = {
        "monday": 0, "mon": 0,
        "tuesday": 1, "tue": 1, "tues": 1,
        "wednesday": 2, "wed": 2,
        "thursday": 3, "thu": 3, "thur": 3, "thurs": 3,
        "friday": 4, "fri": 4,
        "saturday": 5, "sat": 5,
        "sunday": 6, "sun": 6,
    }
    if text in names:
        return names[text]
    for name, number in names.items():
        if re.search(rf"\b{re.escape(name)}\b", text):
            return number
    return None


def _weekday_label(number):
    names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    return names[number] if number is not None and 0 <= number < 7 else ""


def _fill_schedule_guesses(rows):
    """Fill missing weekdays with an editable, chronological training pattern."""
    from calendar_app.services.generation import suggested_weekdays

    for week_number in sorted({row["week"] for row in rows}):
        week_rows = [row for row in rows if row["week"] == week_number]
        day_numbers = sorted({row["day"] for row in week_rows})
        candidates = suggested_weekdays(len(day_numbers)) + list(range(7))
        day_weekdays = {}
        used = set()
        for day_number in day_numbers:
            explicit = next(
                (
                    row.get("weekday")
                    for row in week_rows
                    if row["day"] == day_number and row.get("weekday") is not None
                ),
                None,
            )
            if explicit is not None:
                day_weekdays[day_number] = explicit
                used.add(explicit)
        for day_number in day_numbers:
            if day_number not in day_weekdays:
                weekday = next(candidate for candidate in candidates if candidate not in used)
                day_weekdays[day_number] = weekday
                used.add(weekday)
        for row in week_rows:
            if row.get("weekday") is None:
                row["weekday"] = day_weekdays[row["day"]]
                row["weekday_name"] = _weekday_label(row["weekday"])
                row.setdefault("inferred", []).append("training weekday")
    return rows


def _numeric_weight(value):
    """Read an absolute pound target while ignoring relative/informal loads."""
    text = str(value or "").strip().lower()
    if (
        not text
        or text.startswith("+")
        or not re.search(r"\d", text)
        or "%" in text
        or re.search(r"\b(1rm|tm|rpe|rir)\b", text)
        or (
            re.search(r"[a-z]", text)
            and not re.search(r"\b(lbs?|pounds?|kgs?|kilograms?)\b", text)
        )
    ):
        return None
    match = re.search(r"\d+(?:\.\d+)?", text.replace(",", ""))
    if not match:
        return None
    weight = float(match.group())
    if re.search(r"\b(kgs?|kilograms?)\b", text):
        weight *= 2.2046226218
        weight = round(weight * 2) / 2
    return weight if 0 < weight <= 5000 else None


def _percentage_target(value):
    """Use the first value in either a single percentage or a percentage range."""
    text = str(value or "").strip().replace("%", "")
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None
    percentage = float(match.group())
    return percentage if 0 <= percentage <= 200 else None


def _load_values(value):
    """Interpret one load cell as absolute weight(s) or a percentage."""
    text = str(value or "").strip()
    if not text:
        return None, [], None
    if "%" in text or re.search(
        r"\b(percent|pct|1rm|training max|tm)\b", text, re.IGNORECASE
    ):
        return None, [], _percentage_target(text)
    parts = [part.strip() for part in re.split(r"[,/;|]", text)]
    if len(parts) > 1:
        targets = [_numeric_weight(part) for part in parts]
        if any(weight is not None for weight in targets):
            first = next(weight for weight in targets if weight is not None)
            return first, targets[:30], None
    return _numeric_weight(text), [], None


def _inline_load(value):
    """Read a load written after @, e.g. ``4 x 5 @ 80%``."""
    parts = re.split(r"\s*@\s*", str(value or ""), maxsplit=1)
    return _load_values(parts[1]) if len(parts) == 2 else (None, [], None)


def _rir_from_notes(value):
    match = re.search(r"\bRIR\s*(\d+(?:\.\d+)?)", str(value or ""), re.IGNORECASE)
    if not match:
        return None
    return min(float(match.group(1)), 10)


def _rpe_from_notes(value):
    match = re.search(r"\bRPE\s*(\d+(?:\.\d+)?)", str(value or ""), re.IGNORECASE)
    if not match:
        return None
    return min(float(match.group(1)), 10)


def _direct_effort(value):
    """Read a numeric RIR/RPE cell, including simple ranges such as 1-2."""
    match = re.search(r"\d+(?:\.\d+)?", str(value or ""))
    if not match:
        return None
    effort = float(match.group())
    return effort if 0 <= effort <= 10 else None


def _sets_and_reps(raw_reps, fallback_sets):
    """Split common combined prescriptions such as '3 x 20' or '3 x 15-20'."""
    text = re.split(r"\s*@\s*", str(raw_reps or "").strip(), maxsplit=1)[0]
    match = re.fullmatch(
        r"(\d+)\s*(?:[x×]|sets?\s+(?:of\s+)?)\s*(.+)",
        text,
        re.IGNORECASE,
    )
    if match:
        return min(int(match.group(1)), 30), parse_reps(match.group(2).strip())
    return fallback_sets, parse_reps(text)


def auto_parse_workbook(django_file, extension):
    """Parse multi-tab workout workbooks without asking for manual mappings.

    Handles both flat CSV/XLSX tables and coach-authored layouts where each
    ``Week N`` worksheet has repeated workout sections beginning with an
    Exercise header. It returns an empty list when the structure is not clear,
    leaving the editable manual mapper as the safe fallback.
    """
    if extension == ".csv":
        rows = read_rows(django_file, extension)
        if not rows:
            return []
        mapping = suggest_column_mapping(rows[0])
        if "exercise" not in mapping:
            return []
        parsed, errors = parse_mapped_rows(rows, mapping, has_header=True)
        return [] if errors else parsed
    if extension != ".xlsx":
        return []

    from openpyxl import load_workbook

    django_file.seek(0)
    workbook = load_workbook(django_file, read_only=True, data_only=True)
    parsed = []
    has_named_week_sheets = any(
        re.match(r"^\s*week\s*\d{1,2}\b", sheet.title, re.IGNORECASE)
        for sheet in workbook.worksheets
    )
    parsed_limit = getattr(settings, "MAX_EXCEL_PARSED_ROWS", 5000)
    scanned_limit = max(parsed_limit * 2, MAX_ROWS)
    scanned_rows = 0
    try:
        for sheet in workbook.worksheets:
            rows = []
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= MAX_ROWS:
                    break
                scanned_rows += 1
                if scanned_rows > scanned_limit:
                    raise WorkbookLimitError(
                        f"Workbook contains more than {scanned_limit:,} rows to inspect safely."
                    )
                rows.append([
                    "" if cell is None else str(cell).strip()
                    for cell in row[:MAX_COLS]
                ])
            week = _week_number(sheet.title, rows)
            if week is None and has_named_week_sheets:
                continue

            active = None
            day = 0
            previous_header = -1
            for row_index, row in enumerate(rows):
                exercise_column = next((
                    index for index, label in enumerate(row)
                    if _normalized_label(label) in COLUMN_ALIASES["exercise"]
                ), None)
                if exercise_column is None:
                    mapping = {}
                else:
                    mapping = suggest_column_mapping(row)
                if exercise_column is not None:
                    day += 1
                    title_candidates = []
                    for title_index in range(row_index - 1, previous_header, -1):
                        nonempty = [cell for cell in rows[title_index] if cell]
                        if nonempty:
                            title_candidates.append(str(nonempty[0]))
                    workout_name = next((
                        title for title in title_candidates
                        if re.search(
                            r"\b(upper|lower|push|pull|full body|workout|session|day\s*\d+)\b",
                            title,
                            re.IGNORECASE,
                        )
                        and not re.search(r"\bweek\s*\d+\b", title, re.IGNORECASE)
                    ), title_candidates[0] if title_candidates else "")
                    set_columns = []
                    for column_index, label in enumerate(row):
                        match = re.fullmatch(r"set\s*(\d+)", _normalized_label(label))
                        if match:
                            set_columns.append((int(match.group(1)), column_index))
                    active = {
                        "mapping": mapping,
                        "exercise_column": exercise_column,
                        "set_columns": [column for _number, column in sorted(set_columns)],
                        "workout_name": workout_name or f"Day {day}",
                    }
                    previous_header = row_index
                    continue

                if active is None:
                    continue
                exercise = row[active["exercise_column"]] if active["exercise_column"] < len(row) else ""
                if not str(exercise).strip():
                    active = None
                    continue

                def cell(field):
                    column = active["mapping"].get(field)
                    return row[column] if column is not None and column < len(row) else ""

                set_weights = []
                for column in active["set_columns"]:
                    weight = _numeric_weight(row[column] if column < len(row) else "")
                    # Preserve each original set position. Otherwise a blank
                    # Set 2 would make Set 3's load appear as Set 2's load.
                    set_weights.append(weight)
                first_set_weight = next(
                    (weight for weight in set_weights if weight is not None), None
                )
                if first_set_weight is None:
                    set_weights = []
                fallback_sets = max(len(active["set_columns"]), 1)
                explicit_sets = _int_or_none(cell("sets"), maximum=30)
                target_sets, reps = _sets_and_reps(
                    cell("reps"), explicit_sets or fallback_sets
                )
                notes = str(cell("notes")).strip()
                rir = _direct_effort(cell("rir"))
                if rir is None:
                    rir = _rir_from_notes(notes)
                rpe = _direct_effort(cell("rpe"))
                if rpe is None:
                    rpe = _rpe_from_notes(notes)
                load_weight, load_set_weights, inferred_percentage = _load_values(
                    cell("weight")
                )
                inline_weight, inline_set_weights, inline_percentage = _inline_load(
                    cell("reps")
                )
                if (
                    load_weight is None
                    and not load_set_weights
                    and inferred_percentage is None
                ):
                    load_weight = inline_weight
                    load_set_weights = inline_set_weights
                    inferred_percentage = inline_percentage
                if not set_weights and load_set_weights:
                    set_weights = load_set_weights
                    first_set_weight = load_weight
                    if explicit_sets is None:
                        target_sets = len(load_set_weights)
                elif not set_weights:
                    first_set_weight = load_weight
                percentage = _percentage_target(cell("percentage"))
                inferred = []
                if (inline_weight is not None or inline_set_weights) and not cell("weight"):
                    inferred.append("load from prescription")
                if percentage is None and inferred_percentage is not None:
                    percentage = inferred_percentage
                    inferred.append("percentage from load")
                if percentage is None and "%" in notes:
                    percentage = _percentage_target(notes)
                    if percentage is not None:
                        inferred.append("percentage from notes")
                row_week = _int_or_none(cell("week"), maximum=52) or week or 1
                row_day = _int_or_none(cell("day"), maximum=14) or day
                row_workout = str(cell("workout_name")).strip()
                weekday = _weekday_number(cell("weekday"))
                if weekday is None:
                    weekday = _weekday_number(row_workout or active["workout_name"])
                parsed.append({
                    "row": row_index + 1,
                    "week": row_week,
                    "day": row_day,
                    "weekday": weekday,
                    "weekday_name": _weekday_label(weekday),
                    "workout_name": row_workout[:120] or active["workout_name"],
                    "exercise": str(exercise).strip()[:120],
                    "sets": target_sets,
                    "rep_min": reps["min"],
                    "rep_max": reps["max"],
                    "rep_text": reps["text"],
                    "weight": first_set_weight,
                    "set_weights": set_weights,
                    "percentage": percentage,
                    "rir": rir,
                    "rpe": rpe,
                    "rest_seconds": _parse_rest(cell("rest")),
                    "tempo": str(cell("tempo")).strip()[:20],
                    "notes": notes[:300],
                    "superset": str(cell("superset")).strip()[:5],
                    "inferred": inferred,
                })
                if len(parsed) > parsed_limit:
                    raise WorkbookLimitError(
                        f"Workbook contains more than {parsed_limit:,} exercise rows."
                    )
    finally:
        workbook.close()
    return _fill_schedule_guesses(parsed)


def list_sheets(django_file, extension):
    if extension == ".csv":
        return ["Sheet1"]
    from openpyxl import load_workbook

    django_file.seek(0)
    workbook = load_workbook(django_file, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_rows(django_file, extension, sheet_name=None, limit=MAX_ROWS):
    """Rows as lists of strings, truncated to sane bounds."""
    rows = []
    if extension == ".csv":
        django_file.seek(0)
        text = django_file.read().decode("utf-8-sig", errors="replace")
        for index, row in enumerate(csv.reader(io.StringIO(text))):
            if index >= limit:
                break
            rows.append([str(cell).strip() for cell in row[:MAX_COLS]])
    else:
        from openpyxl import load_workbook

        django_file.seek(0)
        workbook = load_workbook(django_file, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name] if sheet_name else workbook.active
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index >= limit:
                    break
                rows.append([
                    "" if cell is None else str(cell).strip()
                    for cell in row[:MAX_COLS]
                ])
        finally:
            workbook.close()
    return rows


def parse_reps(raw):
    """Parse a rep prescription into {min, max, text}.

    Handles: 5 | 3-5 | 8–12 | 10/leg | AMRAP | 3 rounds | 20-40 yd | 30 seconds
    Unrecognized formats keep the raw text so nothing is lost.
    """
    text = str(raw or "").strip()
    if not text:
        return {"min": None, "max": None, "text": ""}
    plain = text.replace("–", "-").replace("—", "-").lower()
    match = re.fullmatch(r"(\d+)", plain)
    if match:
        value = int(match.group(1))
        return {"min": value, "max": value, "text": ""}
    match = re.fullmatch(r"(\d+)\s*-\s*(\d+)", plain)
    if match:
        low, high = sorted((int(match.group(1)), int(match.group(2))))
        return {"min": low, "max": high, "text": ""}
    match = re.fullmatch(r"(\d+)\s*/\s*(leg|side|arm)", plain)
    if match:
        value = int(match.group(1))
        return {"min": value, "max": value, "text": text}
    # AMRAP, rounds, distances, times, everything else: keep as text
    return {"min": None, "max": None, "text": text[:40]}


def _int_or_none(value, maximum=1000):
    try:
        number = int(float(str(value).strip()))
        return number if 0 < number <= maximum else None
    except (TypeError, ValueError):
        return None


def _float_or_none(value, maximum=5000):
    try:
        number = float(str(value).strip().replace("%", ""))
        return number if 0 <= number <= maximum else None
    except (TypeError, ValueError):
        return None


def _parse_rest(value):
    """Rest as seconds from '90', '90s', '2:00', '2 min'."""
    text = str(value or "").strip().lower()
    if not text:
        return None
    match = re.fullmatch(r"(\d+):(\d{2})", text)
    if match:
        return int(match.group(1)) * 60 + int(match.group(2))
    match = re.fullmatch(r"(\d+(?:\.\d+)?)\s*(min|minutes|m)", text)
    if match:
        return int(float(match.group(1)) * 60)
    match = re.fullmatch(r"(\d+)\s*(s|sec|seconds)?", text)
    if match:
        return int(match.group(1))
    return None


def parse_mapped_rows(rows, mapping, has_header=True):
    """Apply a column mapping ({field: column_index}) to raw rows.

    Returns (parsed_rows, errors). Each parsed row is a plain dict ready to be
    stored in ImportJob.parsed_data.
    """
    parsed, errors = [], []
    start = 1 if has_header else 0
    exercise_column = mapping.get("exercise")
    if exercise_column is None:
        return [], ["An Exercise column mapping is required."]

    def cell(row, field):
        index = mapping.get(field)
        if index is None or index >= len(row):
            return ""
        return row[index]

    for line_number, row in enumerate(rows[start:], start=start + 1):
        exercise_name = str(cell(row, "exercise")).strip()
        if not exercise_name:
            continue
        if _normalized_label(exercise_name) in COLUMN_ALIASES["exercise"]:
            continue
        explicit_sets = _int_or_none(cell(row, "sets"), maximum=30)
        target_sets, reps = _sets_and_reps(
            cell(row, "reps"), explicit_sets or 3
        )
        weight, set_weights, inferred_percentage = _load_values(cell(row, "weight"))
        inline_weight, inline_set_weights, inline_percentage = _inline_load(
            cell(row, "reps")
        )
        if weight is None and not set_weights and inferred_percentage is None:
            weight = inline_weight
            set_weights = inline_set_weights
            inferred_percentage = inline_percentage
        percentage = _percentage_target(cell(row, "percentage"))
        inferred = []
        if (inline_weight is not None or inline_set_weights) and not cell(row, "weight"):
            inferred.append("load from prescription")
        if percentage is None and inferred_percentage is not None:
            percentage = inferred_percentage
            inferred.append("percentage from load")
        parsed.append({
            "row": line_number,
            "week": _int_or_none(cell(row, "week"), maximum=52) or 1,
            "day": _int_or_none(cell(row, "day"), maximum=14) or 1,
            "weekday": _weekday_number(cell(row, "weekday")),
            "weekday_name": _weekday_label(_weekday_number(cell(row, "weekday"))),
            "workout_name": str(cell(row, "workout_name")).strip()[:120],
            "exercise": exercise_name[:120],
            "sets": explicit_sets or (len(set_weights) if set_weights else target_sets),
            "rep_min": reps["min"],
            "rep_max": reps["max"],
            "rep_text": reps["text"],
            "weight": weight,
            "set_weights": set_weights,
            "percentage": percentage,
            "rir": _float_or_none(cell(row, "rir"), maximum=10),
            "rpe": _float_or_none(cell(row, "rpe"), maximum=10),
            "rest_seconds": _parse_rest(cell(row, "rest")),
            "tempo": str(cell(row, "tempo")).strip()[:20],
            "notes": str(cell(row, "notes")).strip()[:300],
            "superset": str(cell(row, "superset")).strip()[:5],
            "inferred": inferred,
        })
    if not parsed:
        errors.append("No rows with an exercise name were found.")
    return _fill_schedule_guesses(parsed), errors
