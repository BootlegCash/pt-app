import io
import zipfile

from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse

from core.tests.utils import link_coach, make_user
from imports.models import ImportJob, ReferenceFile
from imports.services.excel_parser import (
    WorkbookLimitError,
    auto_parse_workbook,
    detect_header_row,
    parse_mapped_rows,
    parse_reps,
    suggest_column_mapping,
)
from imports.services.validation import validate_excel_upload, validate_pdf_upload
from programs.models import Program


def make_xlsx(rows):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Plan"
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


def make_multi_week_xlsx():
    from openpyxl import Workbook

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview & 1RM Tracker"
    overview.append(["PROGRAM OVERVIEW"])
    for week in (1, 2):
        sheet = workbook.create_sheet(f"Week {week}")
        sheet.append([None, f"WEEK {week} - PHASE"])
        sheet.append([None, "Week notes"])
        sheet.append([])
        sheet.append([None, "UPPER A - Chest & Back"])
        sheet.append([None, "Exercise", "Set 1", "Set 2", "Target Reps", "% of 1RM", "Intensity / Notes"])
        sheet.append([None, "Bench Press", "185 lbs", "195 lbs", "4-6", "80-88%", "RIR 1-2"])
        sheet.append([])
        sheet.append([None, "SHOULDERS"])
        sheet.append([None, "Exercise", "Set 1", "Set 2", "Target Reps", "% of 1RM", "Intensity / Notes"])
        sheet.append([None, "Lateral Raise", "20 lbs", "25 lbs", "3 x 15-20", "-", "To failure"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


SAMPLE_ROWS = [
    ["Week", "Day", "Workout", "Exercise", "Sets", "Reps", "Weight", "RIR", "Rest", "Notes"],
    [1, 1, "Upper A", "Bench Press", 4, "6-8", 185, 2, "2:30", "pause first rep"],
    [1, 1, "Upper A", "Chest-Supported Row", 4, "8–12", 115, 2, "90s", ""],
    [1, 2, "Lower A", "Back Squat", 3, "5", 265, 2, "3 min", ""],
    [1, 2, "Lower A", "Walking Lunge", 2, "10/leg", 40, "", "60", ""],
    [2, 1, "Upper A", "Bench Press", 4, "AMRAP", 185, 1, "150", ""],
]


class RepParsingTests(TestCase):
    def test_rep_formats(self):
        self.assertEqual(parse_reps("5"), {"min": 5, "max": 5, "text": ""})
        self.assertEqual(parse_reps("3-5"), {"min": 3, "max": 5, "text": ""})
        self.assertEqual(parse_reps("8–12"), {"min": 8, "max": 12, "text": ""})
        self.assertEqual(parse_reps("10/leg"), {"min": 10, "max": 10, "text": "10/leg"})
        self.assertEqual(parse_reps("AMRAP")["text"], "AMRAP")
        self.assertEqual(parse_reps("3 rounds")["text"], "3 rounds")
        self.assertEqual(parse_reps("20-40 yd")["text"], "20-40 yd")
        self.assertEqual(parse_reps("30 seconds")["text"], "30 seconds")
        self.assertEqual(parse_reps(""), {"min": None, "max": None, "text": ""})

    def test_mapping_requires_exercise_column(self):
        parsed, errors = parse_mapped_rows([["a", "b"]], {"sets": 0})
        self.assertEqual(parsed, [])
        self.assertTrue(errors)

    def test_suggests_common_workout_headers(self):
        mapping = suggest_column_mapping([
            "Week", "Training Day", "Session", "Exercise Name", "Sets",
            "Target Reps", "Load (lb)", "% of 1RM", "RIR", "Rest Time", "Coach Notes",
        ])
        self.assertEqual(mapping, {
            "week": 0, "day": 1, "workout_name": 2, "exercise": 3,
            "sets": 4, "reps": 5, "weight": 6, "percentage": 7,
            "rir": 8, "rest": 9, "notes": 10,
        })

    def test_detects_real_header_below_title_rows(self):
        rows = [
            [None, "RYAN OLSON - 8 WEEK PROGRAM"],
            [None, "WEEK 7 - PEAK"],
            [],
            [None, "UPPER A - Chest & Back"],
            [None, "Exercise", "Set 1", "Set 2", "Target Reps", "% of 1RM", "Notes"],
            [None, "Bench Press", "185 lb", "195 lb", "4-6", "80%", "RIR 1"],
        ]

        header_index, mapping = detect_header_row(rows)

        self.assertEqual(header_index, 4)
        self.assertEqual(mapping["exercise"], 1)
        self.assertEqual(mapping["reps"], 4)
        self.assertEqual(mapping["percentage"], 5)
        self.assertEqual(mapping["notes"], 6)
        self.assertNotIn("sets", mapping)

    def test_does_not_guess_set_weight_column_as_set_count(self):
        mapping = suggest_column_mapping(
            ["Exercise", "Set 1", "Set 2", "Target Reps", "% of 1RM"]
        )

        self.assertNotIn("sets", mapping)

    def test_auto_parses_repeated_workouts_across_week_sheets(self):
        rows = auto_parse_workbook(make_multi_week_xlsx(), ".xlsx")
        self.assertEqual(len(rows), 4)
        self.assertEqual({row["week"] for row in rows}, {1, 2})
        self.assertEqual(rows[0]["workout_name"], "UPPER A - Chest & Back")
        self.assertEqual(rows[0]["exercise"], "Bench Press")
        self.assertEqual(rows[0]["sets"], 2)
        self.assertEqual(rows[0]["set_weights"], [185.0, 195.0])
        self.assertEqual(rows[0]["percentage"], 80.0)
        self.assertEqual(rows[0]["rir"], 1.0)
        self.assertEqual(rows[1]["sets"], 3)
        self.assertEqual(rows[1]["rep_min"], 15)
        self.assertEqual(rows[1]["rep_max"], 20)

    def test_auto_parser_preserves_blank_set_weight_positions(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Week 1"
        sheet.append([None, "UPPER A"])
        sheet.append([
            None, "Exercise", "Set 1", "Set 2", "Set 3", "Target Reps", "RIR", "RPE",
        ])
        sheet.append([
            None, "Bench Press", "185 lbs", None, "205 lbs", "3 x 5", "2", "8.5",
        ])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        rows = auto_parse_workbook(buffer, ".xlsx")

        self.assertEqual(rows[0]["set_weights"], [185.0, None, 205.0])
        self.assertEqual(rows[0]["weight"], 185.0)
        self.assertEqual(rows[0]["rir"], 2.0)
        self.assertEqual(rows[0]["rpe"], 8.5)

    def test_auto_parser_understands_flat_sets_and_percentage_load(self):
        workbook = make_xlsx([
            ["Week", "Day", "Workout", "Exercise", "Sets", "Reps", "Load"],
            [1, 2, "Lower", "Back Squat", 4, 5, "80% 1RM"],
        ])

        rows = auto_parse_workbook(io.BytesIO(workbook), ".xlsx")

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["sets"], 4)
        self.assertEqual(rows[0]["rep_min"], 5)
        self.assertIsNone(rows[0]["weight"])
        self.assertEqual(rows[0]["percentage"], 80.0)
        self.assertIn("percentage from load", rows[0]["inferred"])

    def test_manual_mapping_treats_percent_load_as_percentage(self):
        parsed, errors = parse_mapped_rows(
            [["Exercise", "Sets", "Reps", "Load"], ["Bench", 4, 5, "75%"]],
            {"exercise": 0, "sets": 1, "reps": 2, "weight": 3},
        )

        self.assertFalse(errors)
        self.assertEqual(parsed[0]["sets"], 4)
        self.assertIsNone(parsed[0]["weight"])
        self.assertEqual(parsed[0]["percentage"], 75.0)

    def test_guesses_combined_prescription_column(self):
        header = ["Exercise", "Prescription"]
        mapping = suggest_column_mapping(header)
        self.assertEqual(mapping, {"exercise": 0, "reps": 1})

        parsed, errors = parse_mapped_rows(
            [header, ["Bench", "4 x 5 @ 80%"]], mapping
        )

        self.assertFalse(errors)
        self.assertEqual(parsed[0]["sets"], 4)
        self.assertEqual(parsed[0]["rep_min"], 5)
        self.assertEqual(parsed[0]["percentage"], 80.0)

    def test_auto_interprets_a_flat_csv(self):
        rows = auto_parse_workbook(
            io.BytesIO(b"Exercise,Sets,Reps,Load\nBench Press,4,5,80%\n"),
            ".csv",
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["sets"], 4)
        self.assertEqual(rows[0]["percentage"], 80.0)

    @override_settings(MAX_EXCEL_PARSED_ROWS=1)
    def test_auto_parser_rejects_too_many_exercise_rows(self):
        workbook = make_xlsx([
            ["Exercise", "Sets", "Reps"],
            ["Bench", 3, 5],
            ["Squat", 3, 5],
        ])

        with self.assertRaises(WorkbookLimitError):
            auto_parse_workbook(io.BytesIO(workbook), ".xlsx")


class ValidationTests(TestCase):
    def test_rejects_wrong_extension(self):
        upload = SimpleUploadedFile("evil.exe", b"MZ....")
        with self.assertRaises(ValidationError):
            validate_excel_upload(upload)

    def test_rejects_renamed_binary_as_xlsx(self):
        upload = SimpleUploadedFile("fake.xlsx", b"not a zip file")
        with self.assertRaises(ValidationError):
            validate_excel_upload(upload)

    @override_settings(MAX_EXCEL_UPLOAD_MB=1)
    def test_rejects_oversize(self):
        upload = SimpleUploadedFile("big.csv", b"a," * 800_000)
        with self.assertRaises(ValidationError):
            validate_excel_upload(upload)

    def test_accepts_real_xlsx_and_csv(self):
        xlsx = SimpleUploadedFile("plan.xlsx", make_xlsx(SAMPLE_ROWS))
        self.assertEqual(validate_excel_upload(xlsx), ".xlsx")
        csv = SimpleUploadedFile("plan.csv", b"Exercise,Sets\nBench,3\n")
        self.assertEqual(validate_excel_upload(csv), ".csv")

    @override_settings(MAX_EXCEL_UNCOMPRESSED_MB=1)
    def test_rejects_small_compressed_workbook_that_expands_too_large(self):
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", "types")
            archive.writestr("xl/workbook.xml", "workbook")
            archive.writestr("xl/worksheets/sheet1.xml", "x" * (2 * 1024 * 1024))
        upload = SimpleUploadedFile("large-expanded.xlsx", buffer.getvalue())

        with self.assertRaisesMessage(ValidationError, "safe processing limit"):
            validate_excel_upload(upload)

    def test_pdf_validation(self):
        good = SimpleUploadedFile("doc.pdf", b"%PDF-1.7 fake body")
        self.assertEqual(validate_pdf_upload(good), ".pdf")
        bad = SimpleUploadedFile("doc.pdf", b"<html>nope</html>")
        with self.assertRaises(ValidationError):
            validate_pdf_upload(bad)


class ImportWorkflowTests(TestCase):
    def setUp(self):
        self.coach = make_user(is_coach=True)
        self.athlete = make_user()
        link_coach(self.coach, self.athlete)
        self.client.force_login(self.athlete)

    def _upload(self):
        upload = SimpleUploadedFile("plan.xlsx", make_xlsx(SAMPLE_ROWS))
        self.client.post(reverse("imports:upload_excel"), {"file": upload})
        return ImportJob.objects.get(user=self.athlete)

    def _map_and_submit(self, job):
        self.client.post(reverse("imports:select_sheet", args=[job.uuid]), {"sheet": "Plan"})
        self.client.post(reverse("imports:mapping", args=[job.uuid]), {
            "map_week": "0", "map_day": "1", "map_workout_name": "2",
            "map_exercise": "3", "map_sets": "4", "map_reps": "5",
            "map_weight": "6", "map_rir": "7", "map_rest": "8", "map_notes": "9",
            "has_header": "on",
        })
        self.client.post(reverse("imports:review_parsed", args=[job.uuid]))
        job.refresh_from_db()
        return job

    def test_full_flow_to_submission(self):
        job = self._map_and_submit(self._upload())
        self.assertEqual(job.status, ImportJob.Status.SUBMITTED)
        self.assertEqual(len(job.parsed_data), 5)
        first = job.parsed_data[0]
        self.assertEqual(first["exercise"], "Bench Press")
        self.assertEqual(first["rep_min"], 6)
        self.assertEqual(first["rest_seconds"], 150)

    def test_source_workbook_download_is_not_cacheable(self):
        job = self._upload()

        response = self.client.get(
            reverse("imports:download_import_file", args=[job.uuid])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Cache-Control"], "private, no-store")
        response.close()

    def test_approval_creates_draft_program_not_assignment(self):
        job = self._map_and_submit(self._upload())
        self.client.force_login(self.coach)
        response = self.client.post(
            reverse("imports:approve_job", args=[job.uuid]), {"notes": "looks good"}
        )
        self.assertEqual(response.status_code, 302)
        job.refresh_from_db()
        self.assertEqual(job.status, ImportJob.Status.IMPORTED)
        program = job.created_program
        self.assertIsNotNone(program)
        self.assertEqual(program.status, Program.Status.DRAFT)
        self.assertIsNone(program.assigned_to)  # never auto-assigned
        self.assertEqual(program.number_of_weeks, 2)
        self.assertEqual(
            sum(w.days.count() for w in program.weeks.all()), 3
        )

    def test_import_never_reuses_another_coachs_private_exercise(self):
        from exercises.models import Exercise

        other_coach = make_user(is_coach=True)
        private = Exercise.objects.create(
            name="Bench Press",
            primary_muscle=Exercise.Muscle.CHEST,
            public=False,
            created_by=other_coach,
            coaching_cues="Other coach's private cues",
        )
        job = self._map_and_submit(self._upload())
        self.client.force_login(self.coach)

        self.client.post(reverse("imports:approve_job", args=[job.uuid]))

        job.refresh_from_db()
        imported = (
            job.created_program.weeks.first().days.first().exercises.first().exercise
        )
        self.assertNotEqual(imported, private)
        self.assertEqual(imported.created_by, self.coach)
        self.assertFalse(imported.public)
        self.assertNotEqual(imported.coaching_cues, private.coaching_cues)

    def test_coach_can_upload_and_create_a_draft_directly(self):
        self.client.force_login(self.coach)
        upload = SimpleUploadedFile("coach-plan.xlsx", make_xlsx(SAMPLE_ROWS))
        response = self.client.post(reverse("imports:coach_upload"), {"file": upload})
        self.assertEqual(response.status_code, 302)
        job = ImportJob.objects.get(user=self.coach)
        self.assertRedirects(
            response, reverse("imports:review_parsed", args=[job.uuid])
        )
        response = self.client.post(reverse("imports:review_parsed", args=[job.uuid]))

        self.assertEqual(response.status_code, 302)
        job.refresh_from_db()
        self.assertEqual(job.status, ImportJob.Status.IMPORTED)
        self.assertEqual(job.created_program.owner, self.coach)
        self.assertEqual(job.created_program.status, Program.Status.DRAFT)
        self.assertIsNone(job.created_program.assigned_to)
        first = job.created_program.weeks.first().days.first().exercises.first()
        self.assertEqual(first.progression_method, "manual")

    def test_coach_multi_week_upload_skips_manual_mapping(self):
        self.client.force_login(self.coach)
        upload = SimpleUploadedFile("multi-week.xlsx", make_multi_week_xlsx().read())
        response = self.client.post(reverse("imports:coach_upload"), {"file": upload})

        job = ImportJob.objects.get(user=self.coach)
        self.assertRedirects(response, reverse("imports:review_parsed", args=[job.uuid]))
        self.assertEqual(job.mapping_configuration["mode"], "automatic_multi_sheet")
        self.assertEqual(len(job.parsed_data), 4)

        revisit = self.client.get(reverse("imports:mapping", args=[job.uuid]))
        self.assertRedirects(revisit, reverse("imports:review_parsed", args=[job.uuid]))
        job.refresh_from_db()
        self.assertEqual(job.status, ImportJob.Status.MAPPING)

        manual = self.client.get(
            reverse("imports:mapping", args=[job.uuid]) + "?manual=1"
        )
        self.assertRedirects(manual, reverse("imports:select_sheet", args=[job.uuid]))
        job.refresh_from_db()
        self.assertEqual(job.parsed_data, [])
        self.assertEqual(job.mapping_configuration, {"mode": "manual"})

        self.client.post(
            reverse("imports:select_sheet", args=[job.uuid]), {"sheet": "Week 1"}
        )
        mapping_page = self.client.get(reverse("imports:mapping", args=[job.uuid]))
        self.assertEqual(mapping_page.status_code, 200)
        self.assertTemplateUsed(mapping_page, "imports/mapping.html")

    def test_review_paginates_auto_detected_rows(self):
        self.client.force_login(self.coach)
        upload = SimpleUploadedFile("multi-week.xlsx", make_multi_week_xlsx().read())
        self.client.post(reverse("imports:coach_upload"), {"file": upload})
        job = ImportJob.objects.get(user=self.coach)
        job.parsed_data = job.parsed_data * 30
        job.save(update_fields=["parsed_data"])

        response = self.client.get(reverse("imports:review_parsed", args=[job.uuid]))

        self.assertEqual(len(response.context["rows"]), 100)
        self.assertEqual(response.context["total"], 120)
        page_two = self.client.get(
            reverse("imports:review_parsed", args=[job.uuid]) + "?page=2"
        )
        self.assertEqual(len(page_two.context["rows"]), 20)

    def test_repeated_create_request_does_not_duplicate_program(self):
        self.client.force_login(self.coach)
        upload = SimpleUploadedFile("coach-plan.xlsx", make_xlsx(SAMPLE_ROWS))
        self.client.post(reverse("imports:coach_upload"), {"file": upload})
        job = ImportJob.objects.get(user=self.coach)

        self.client.post(reverse("imports:review_parsed", args=[job.uuid]))
        first_program_id = ImportJob.objects.get(pk=job.pk).created_program_id
        self.client.post(reverse("imports:review_parsed", args=[job.uuid]))

        job.refresh_from_db()
        self.assertEqual(job.created_program_id, first_program_id)
        self.assertEqual(Program.objects.filter(source_imports=job).count(), 1)

    def test_rejection(self):
        job = self._map_and_submit(self._upload())
        self.client.force_login(self.coach)
        self.client.post(reverse("imports:reject_job", args=[job.uuid]), {"notes": "wrong file"})
        job.refresh_from_db()
        self.assertEqual(job.status, ImportJob.Status.REJECTED)
        self.assertIsNone(job.created_program)

    def test_client_cannot_approve_own_upload(self):
        job = self._map_and_submit(self._upload())
        response = self.client.post(reverse("imports:approve_job", args=[job.uuid]))
        self.assertEqual(response.status_code, 403)

    def test_unrelated_coach_cannot_approve(self):
        job = self._map_and_submit(self._upload())
        stranger_coach = make_user(is_coach=True)
        self.client.force_login(stranger_coach)
        response = self.client.post(reverse("imports:approve_job", args=[job.uuid]))
        self.assertEqual(response.status_code, 403)


class PrivateFileAccessTests(TestCase):
    def setUp(self):
        self.coach = make_user(is_coach=True)
        self.athlete = make_user()
        self.stranger = make_user()
        link_coach(self.coach, self.athlete)
        self.client.force_login(self.athlete)
        self.client.post(reverse("imports:upload_pdf"), {
            "file": SimpleUploadedFile("notes.pdf", b"%PDF-1.7 body"),
        })
        self.pdf = ReferenceFile.objects.get(user=self.athlete)
        self.url = reverse("imports:download_pdf", args=[self.pdf.uuid])

    def test_owner_and_coach_can_download(self):
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Disposition"], 'attachment; filename="notes.pdf"'
        )
        self.assertEqual(response["Cache-Control"], "private, no-store")
        response.close()
        self.client.force_login(self.coach)
        response = self.client.get(self.url)
        self.assertEqual(response.status_code, 200)
        response.close()

    def test_stranger_and_anonymous_cannot(self):
        self.client.force_login(self.stranger)
        self.assertEqual(self.client.get(self.url).status_code, 404)
        self.client.logout()
        self.assertEqual(self.client.get(self.url).status_code, 302)  # to login

    def test_stored_filename_randomized(self):
        self.assertNotIn("notes", self.pdf.file.name)
        self.assertEqual(self.pdf.original_filename, "notes.pdf")

    def test_client_cannot_open_coach_file_management_screen(self):
        response = self.client.get(
            reverse("imports:client_files", args=[self.athlete.uuid])
        )
        self.assertEqual(response.status_code, 403)

    def test_assigned_coach_can_manage_pdf_notes(self):
        self.client.force_login(self.coach)
        list_url = reverse("imports:client_files", args=[self.athlete.uuid])
        self.assertEqual(self.client.get(list_url).status_code, 200)
        edit_url = reverse("imports:pdf_edit", args=[self.pdf.uuid])
        response = self.client.post(edit_url, {"coach_notes": "Review at check-in"})
        self.assertRedirects(response, list_url)
        self.pdf.refresh_from_db()
        self.assertEqual(self.pdf.coach_notes, "Review at check-in")

    def test_coach_notes_are_hidden_from_client_page_and_data_export(self):
        from core.services.export import export_user_data

        self.pdf.coach_notes = "Private coach-only context"
        self.pdf.save(update_fields=["coach_notes"])

        response = self.client.get(reverse("imports:my_files"))
        self.assertNotContains(response, self.pdf.coach_notes)
        exported = export_user_data(self.athlete)
        self.assertNotIn("coach_notes", exported["reference_files"][0])

    def test_dual_role_client_still_cannot_see_their_coachs_private_file_notes(self):
        self.athlete.is_coach = True
        self.athlete.save(update_fields=["is_coach"])
        self.pdf.coach_notes = "Private coach-only context"
        self.pdf.save(update_fields=["coach_notes"])

        response = self.client.get(reverse("imports:my_files"))

        self.assertNotContains(response, self.pdf.coach_notes)

    def test_pdf_cannot_be_linked_to_another_clients_program(self):
        other_athlete = make_user()
        link_coach(self.coach, other_athlete)
        other_program = Program.objects.create(
            owner=self.coach, assigned_to=other_athlete, name="Other plan"
        )
        self.client.force_login(self.coach)
        response = self.client.post(
            reverse("imports:pdf_edit", args=[self.pdf.uuid]),
            {"coach_notes": "private", "program": other_program.pk},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Select a valid choice")
        self.pdf.refresh_from_db()
        self.assertIsNone(self.pdf.program)
